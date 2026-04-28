from flask import Flask, render_template, request, jsonify, Response, stream_with_context, session, redirect, url_for, make_response
from functools import wraps
from google_places_client import GooglePlacesClient
from apollo_client import ApolloClient
from supabase_client import SupabaseClient
from config import Config
import json
import time
import threading
from datetime import datetime
import os
import logging
import requests
import re

# Setup logger
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.config.from_object(Config)
app.secret_key = Config.SECRET_KEY

# Production settings for Vercel/serverless
app.config['SESSION_COOKIE_SECURE'] = os.getenv('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours

# Initialize clients
google_client = GooglePlacesClient()
apollo_client = ApolloClient()

# Initialize Supabase client (ONLY database - Google Sheets removed)
# Initialize lazily to avoid Vercel cold start issues
supabase_client = None

def _build_search_queries(industry: str, location_label: str) -> list:
    """
    Build diverse search queries that work for ANY industry — no hardcoded synonym map.
    Uses multiple phrasing patterns so Google returns different result sets each time.
    """
    queries = []
    ind = industry.strip() if industry else ''

    if ind:
        # --- Pattern 1: Direct industry + location combos ---
        queries.append(f"{ind} in {location_label}")
        queries.append(f"{ind} company {location_label}")
        queries.append(f"{ind} near {location_label}")
        # --- Pattern 2: Generic business terms + location ---
        queries.append(f"{ind} industry {location_label}")
        queries.append(f"{ind} pvt ltd {location_label}")
        queries.append(f"{ind} factory {location_label}")
        queries.append(f"{ind} services {location_label}")
        # --- Pattern 3: Broader terms to catch unlabeled businesses ---
        queries.append(f"companies in {location_label}")
        queries.append(f"industries in {location_label}")
        queries.append(f"businesses in {location_label}")
    else:
        queries.append(f"businesses in {location_label}")
        queries.append(f"companies in {location_label}")
        queries.append(f"industries in {location_label}")
        queries.append(f"pvt ltd in {location_label}")

    return queries


def _geocode_location(address: str) -> tuple:
    """Geocode an address/PIN to (lat, lng) or (None, None) on failure."""
    import requests
    geocode_url = "https://maps.googleapis.com/maps/api/geocode/json"
    geocode_params = {'address': address, 'key': google_client.api_key}
    try:
        geo_resp = requests.get(geocode_url, params=geocode_params, timeout=10)
        geo_data = geo_resp.json()
        if geo_data.get('status') == 'OK' and geo_data.get('results'):
            loc = geo_data['results'][0]['geometry']['location']
            logger.info(f"Geocoded '{address}' → lat={loc['lat']}, lng={loc['lng']}")
            return loc['lat'], loc['lng']
    except Exception as geo_err:
        logger.warning(f"Geocoding error for '{address}': {geo_err}")
    return None, None


def _text_search_page(query: str, lat, lng, radius: int, page_token=None) -> dict:
    """Execute one page of Google Places Text Search."""
    import requests
    params = {
        'query': query,
        'key': google_client.api_key,
    }
    if lat is not None and lng is not None:
        params['location'] = f"{lat},{lng}"
        params['radius'] = radius
    if page_token:
        params['pagetoken'] = page_token
        time.sleep(3)  # Google requires delay before using next_page_token

    resp = requests.get(f"{google_client.base_url}/textsearch/json", params=params, timeout=15)
    return resp.json()


def _nearby_search_page(lat, lng, radius: int, keyword: str = None, page_token=None) -> dict:
    """Execute one page of Google Places Nearby Search (different result set than text search)."""
    import requests
    params = {
        'location': f"{lat},{lng}",
        'radius': radius,
        'key': google_client.api_key,
    }
    if keyword:
        params['keyword'] = keyword
    if page_token:
        params['pagetoken'] = page_token
        time.sleep(3)

    resp = requests.get(f"{google_client.base_url}/nearbysearch/json", params=params, timeout=15)
    return resp.json()


def _fetch_and_yield_places(places_list: list, seen_place_ids: set, industry: str,
                            location_label: str, location_type: str) -> tuple:
    """
    Fetch details for a list of raw place results. Yields (company_dict) for each valid one.
    Returns number of new companies found.
    """
    count = 0
    for place in places_list:
        place_id = place.get('place_id')
        if not place_id or place_id in seen_place_ids:
            continue
        seen_place_ids.add(place_id)

        # Retry up to 2 times
        details = None
        for attempt in range(2):
            details = google_client.get_place_details(place_id)
            if details:
                break
            time.sleep(0.5)

        if details:
            details['place_type'] = details.get('industry', '')
            details['industry'] = industry.strip() if industry else details.get('industry', '')
            if location_type == 'pin':
                details['pin_code'] = location_label
            else:
                details['search_location'] = location_label
                details['place_name'] = location_label
            count += 1
            yield details
        time.sleep(0.1)


def _run_text_search_all_pages(query, lat, lng, radius, seen_place_ids, industry,
                                location_label, location_type, max_results, companies_found):
    """Run text search with full pagination (up to 3 pages = 60 results)."""
    next_page_token = None
    page = 1
    found = 0
    invalid_token_retries = 0

    while companies_found + found < max_results:
        data = _text_search_page(query, lat, lng, radius, next_page_token)
        status = data.get('status', 'UNKNOWN')

        if status != 'OK':
            # Google page tokens can return INVALID_REQUEST for a short time
            # before the next page becomes available.
            if status == 'INVALID_REQUEST' and next_page_token and invalid_token_retries < 3:
                invalid_token_retries += 1
                time.sleep(2)
                continue
            if status == 'OVER_QUERY_LIMIT':
                logger.warning(f"API quota exceeded on query '{query}'")
            break
        invalid_token_retries = 0

        results = data.get('results', [])
        logger.info(f"TextSearch '{query}' r={radius} p{page}: {len(results)} raw results")

        for item in _fetch_and_yield_places(results, seen_place_ids, industry,
                                             location_label, location_type):
            found += 1
            yield item
            if companies_found + found >= max_results:
                return

        next_page_token = data.get('next_page_token')
        if not next_page_token:
            break
        page += 1

    return


def _run_nearby_search_all_pages(keyword, lat, lng, radius, seen_place_ids, industry,
                                  location_label, location_type, max_results, companies_found):
    """Run Nearby Search with full pagination."""
    next_page_token = None
    page = 1
    found = 0
    invalid_token_retries = 0

    while companies_found + found < max_results:
        data = _nearby_search_page(lat, lng, radius, keyword, next_page_token)
        status = data.get('status', 'UNKNOWN')

        if status != 'OK':
            if status == 'INVALID_REQUEST' and next_page_token and invalid_token_retries < 3:
                invalid_token_retries += 1
                time.sleep(2)
                continue
            break
        invalid_token_retries = 0

        results = data.get('results', [])
        logger.info(f"NearbySearch kw='{keyword}' r={radius} p{page}: {len(results)} raw results")

        for item in _fetch_and_yield_places(results, seen_place_ids, industry,
                                             location_label, location_type):
            found += 1
            yield item
            if companies_found + found >= max_results:
                return

        next_page_token = data.get('next_page_token')
        if not next_page_token:
            break
        page += 1

    return


def search_places_progressively(place_name: str, industry: str, max_results: int,
                                 place_idx: int = 1, total_places: int = 1):
    """
    Multi-strategy search for place names. Uses:
    1. Text Search with diverse queries
    2. Nearby Search API (returns different results)
    3. Progressive radius expansion (25km → 50km)
    """
    try:
        lat, lng = _geocode_location(place_name)
        seen_place_ids = set()
        companies_found = 0

        queries = _build_search_queries(industry, place_name)

        # --- Strategy 1: Text Search across multiple queries at 50km ---
        for query in queries:
            if companies_found >= max_results:
                break
            for company in _run_text_search_all_pages(
                query, lat, lng, 50000, seen_place_ids, industry,
                place_name, 'place', max_results, companies_found
            ):
                companies_found += 1
                yield company
                if companies_found >= max_results:
                    break

        # --- Strategy 2: Nearby Search (different API = different results) ---
        if companies_found < max_results and lat is not None:
            nearby_keywords = [industry] if industry else ['company']
            if industry:
                nearby_keywords.extend([f"{industry} company", f"{industry} business", f"{industry} services"])

            for kw in nearby_keywords:
                if companies_found >= max_results:
                    break
                for company in _run_nearby_search_all_pages(
                    kw, lat, lng, 50000, seen_place_ids, industry,
                    place_name, 'place', max_results, companies_found
                ):
                    companies_found += 1
                    yield company
                    if companies_found >= max_results:
                        break

        # --- Strategy 3: Grid search — offset coordinates to cover surrounding areas ---
        if companies_found < max_results and lat is not None:
            offset = 0.12  # ~13km offset
            grid_points = [
                (lat + offset, lng),
                (lat - offset, lng),
                (lat, lng + offset),
                (lat, lng - offset),
            ]
            grid_query = f"{industry} in {place_name}" if industry else f"businesses in {place_name}"

            for g_lat, g_lng in grid_points:
                if companies_found >= max_results:
                    break
                for company in _run_text_search_all_pages(
                    grid_query, g_lat, g_lng, 25000, seen_place_ids, industry,
                    place_name, 'place', max_results, companies_found
                ):
                    companies_found += 1
                    yield company
                    if companies_found >= max_results:
                        break

        logger.info(f"Place '{place_name}': total {companies_found} companies found across all strategies")

    except Exception as e:
        raise Exception(f"Google Places API error for place {place_name}: {str(e)}") from e


def search_pins_progressively(pin_code: str, industry: str, max_results: int,
                               pin_idx: int = 1, total_pins: int = 1):
    """
    Multi-strategy search for PIN codes. Uses:
    1. Text Search with diverse queries
    2. Nearby Search API (returns different results)
    3. Progressive radius expansion (10km → 20km → 35km)
    Designed to work for ANY industry, not just manufacturing.
    """
    try:
        lat, lng = _geocode_location(pin_code)
        seen_place_ids = set()
        companies_found = 0

        queries = _build_search_queries(industry, pin_code)

        # --- Strategy 1: Text Search — progressive radius expansion ---
        radii = [10000, 20000, 35000]  # 10km → 20km → 35km
        for radius in radii:
            if companies_found >= max_results:
                break
            for query in queries:
                if companies_found >= max_results:
                    break
                for company in _run_text_search_all_pages(
                    query, lat, lng, radius, seen_place_ids, industry,
                    pin_code, 'pin', max_results, companies_found
                ):
                    companies_found += 1
                    yield company
                    if companies_found >= max_results:
                        break

            # If we already have enough from smaller radius, no need to expand
            if companies_found >= max_results * 0.8:
                break

        # --- Strategy 2: Nearby Search API (completely different endpoint & result set) ---
        if companies_found < max_results and lat is not None:
            nearby_keywords = [industry] if industry else ['company']
            if industry:
                nearby_keywords.extend([f"{industry} company", f"{industry} business", f"{industry} services"])

            for radius in [15000, 30000]:
                if companies_found >= max_results:
                    break
                for kw in nearby_keywords:
                    if companies_found >= max_results:
                        break
                    for company in _run_nearby_search_all_pages(
                        kw, lat, lng, radius, seen_place_ids, industry,
                        pin_code, 'pin', max_results, companies_found
                    ):
                        companies_found += 1
                        yield company
                        if companies_found >= max_results:
                            break

        # --- Strategy 3: Grid search — offset coordinates N/S/E/W to cover blind spots ---
        if companies_found < max_results and lat is not None:
            offset = 0.08  # ~8-9km offset
            grid_points = [
                (lat + offset, lng),          # North
                (lat - offset, lng),          # South
                (lat, lng + offset),          # East
                (lat, lng - offset),          # West
            ]
            grid_query = f"{industry} in {pin_code}" if industry else f"businesses in {pin_code}"

            for g_lat, g_lng in grid_points:
                if companies_found >= max_results:
                    break
                for company in _run_text_search_all_pages(
                    grid_query, g_lat, g_lng, 15000, seen_place_ids, industry,
                    pin_code, 'pin', max_results, companies_found
                ):
                    companies_found += 1
                    yield company
                    if companies_found >= max_results:
                        break

        logger.info(f"PIN {pin_code}: total {companies_found} companies found across all strategies")

    except Exception as e:
        raise Exception(f"Google Places API error for PIN {pin_code}: {str(e)}") from e

def get_supabase_client():
    """Lazy initialization of Supabase client"""
    global supabase_client
    if supabase_client is None:
        try:
            supabase_client = SupabaseClient()
            print("✅ Using Supabase as backend database")
        except Exception as e:
            print(f"❌ Supabase client not initialized: {str(e)}")
            print("❌ Please check your Supabase configuration in config.py")
            print("❌ Make sure SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are set")
            print("❌ Also ensure you've run the SQL schema in Supabase (see supabase_schema.sql)")
            raise  # Re-raise for actual usage
    return supabase_client

def _company_fingerprint(company):
    """Normalized (name, address) for duplicate detection when place_id differs or is missing."""
    name = (company.get('company_name') or '').strip().lower()
    addr = (company.get('address') or '').strip().lower()
    name_slug = ' '.join(name.split())[:45] if name else ''
    addr_prefix = ' '.join(addr.split())[:40] if addr else ''
    return (name_slug, addr_prefix)

import re

def _is_same_company_by_name_address(company, existing_list, _seen_fingerprints=None):
    """Deduplication via place_id and fingerprint is sufficient. Name-based matching
    was too aggressive (splitting on dashes/commas destroyed distinct names).
    This now always returns False — kept as a no-op so callers don't break."""
    return False

def _normalize_employee_ranges(data):
    """Normalize employee_ranges from request data to a list (empty = no filter)."""
    employee_ranges = data.get('employee_ranges', [])
    if employee_ranges is None:
        employee_ranges = []
    elif isinstance(employee_ranges, (int, float)):
        employee_ranges = []
    elif isinstance(employee_ranges, str):
        employee_ranges = [employee_ranges] if employee_ranges and employee_ranges.lower() != 'all' else []
    elif isinstance(employee_ranges, tuple):
        employee_ranges = list(employee_ranges)
    elif not isinstance(employee_ranges, list):
        employee_ranges = []
    if not employee_ranges and data.get('employee_range'):
        s = (data.get('employee_range') or '').strip()
        employee_ranges = [s] if s and s.lower() != 'all' else []
    return employee_ranges

def filter_companies_by_employee_range(companies, employee_ranges):
    """
    Filter companies by employee range(s).
    employee_ranges: List of ranges like ["50-100", "100-250"] or single string for backward compatibility
    """
    try:
        if employee_ranges is None:
            employee_ranges = []
        elif isinstance(employee_ranges, (int, float)):
            # If it's a number, treat as no filter (empty list)
            employee_ranges = []
        elif isinstance(employee_ranges, str):
            # Handle backward compatibility (single string)
            employee_ranges = [employee_ranges] if employee_ranges and employee_ranges.lower() != 'all' else []
        elif isinstance(employee_ranges, tuple):
            # Convert tuple to list
            employee_ranges = list(employee_ranges)
        elif not isinstance(employee_ranges, list):
            # If it's any other type that's not a list, convert to empty list
            employee_ranges = []
    except Exception as e:
        logger.warning(f"employee_ranges validation failed: {e}, defaulting to no filter")
        employee_ranges = []

    if not employee_ranges or len(employee_ranges) == 0:
        return companies
    
    filtered = []
    
    for company in companies:
        total_employees_str = company.get('total_employees', '') or ''
        if not total_employees_str:
            # If company doesn't have employee count, skip it when filtering
            continue
        
        # Try to extract numeric value from employee string (e.g., "50-100" -> 75, "500+" -> 500)
        employee_count = None
        try:
            # Clean the string first
            cleaned = total_employees_str.strip().replace(',', '').replace(' ', '')
            
            # Validate: reject if contains non-numeric characters (except - and +)
            if not cleaned or not all(c.isdigit() or c in ['-', '+'] for c in cleaned):
                continue
            
            # Handle ranges like "50-100" -> take midpoint
            if '-' in cleaned:
                parts = cleaned.split('-')
                if len(parts) == 2:
                    low = int(parts[0])
                    high = int(parts[1])
                    # Validate range is reasonable
                    if low > 0 and high > low and high <= 1000000:
                        employee_count = (low + high) // 2
            # Handle "500+" or "5000+"
            elif '+' in cleaned:
                num = int(cleaned.replace('+', ''))
                # Validate number is reasonable
                if num > 0 and num <= 1000000:
                    employee_count = num
            # Handle single number
            else:
                num = int(cleaned)
                # Validate number is reasonable
                if num > 0 and num <= 1000000:
                    employee_count = num
        except (ValueError, AttributeError):
            # If we can't parse, skip this company when filtering
            continue
        
        if employee_count is None:
            continue
        
        # Check if company matches ANY of the selected ranges
        matches = False
        # Extra safety: ensure employee_ranges is iterable before looping
        try:
            iter(employee_ranges)
        except TypeError:
            # If not iterable, skip this company (safety fallback)
            continue
        
        for employee_range in employee_ranges:
            if employee_range == "1-10":
                matches = 1 <= employee_count <= 10
            elif employee_range == "10-50":
                matches = 10 <= employee_count <= 50
            elif employee_range == "50-100":
                matches = 50 <= employee_count <= 100
            elif employee_range == "100-250":
                matches = 100 <= employee_count <= 250
            elif employee_range == "250-500":
                matches = 250 <= employee_count <= 500
            elif employee_range == "500-1000":
                matches = 500 <= employee_count <= 1000
            elif employee_range == "1000-5000":
                matches = 1000 <= employee_count <= 5000
            elif employee_range == "5000+":
                matches = employee_count >= 5000
            
            if matches:
                break  # Company matches at least one range, no need to check others
        
        if matches:
            filtered.append(company)
    
    return filtered

# Try to initialize on startup (but don't crash the app if it fails)
try:
    supabase_client = SupabaseClient()
    print("✅ Using Supabase as backend database")
except Exception as e:
    print(f"⚠️  Supabase client initialization deferred: {str(e)}")
    # Will be initialized on first use via get_supabase_client()

# Progress tracking is now handled by Supabase (see get_supabase_client().py)
# This keeps progress persistent across serverless invocations

# Indian states list
INDIAN_STATES = [
    'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Chhattisgarh',
    'Goa', 'Gujarat', 'Haryana', 'Himachal Pradesh', 'Jharkhand',
    'Karnataka', 'Kerala', 'Madhya Pradesh', 'Maharashtra', 'Manipur',
    'Meghalaya', 'Mizoram', 'Nagaland', 'Odisha', 'Punjab',
    'Rajasthan', 'Sikkim', 'Tamil Nadu', 'Telangana', 'Tripura',
    'Uttar Pradesh', 'Uttarakhand', 'West Bengal',
    'Andaman and Nicobar Islands', 'Chandigarh', 'Dadra and Nagar Haveli',
    'Daman and Diu', 'Delhi', 'Jammu and Kashmir', 'Ladakh', 'Lakshadweep', 'Puducherry'
]

# Authentication decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Login route
@app.route('/login')
def login():
    """Login page"""
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')

# Login API endpoint
@app.route('/api/login', methods=['POST'])
def api_login():
    """Handle login authentication"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        # Check credentials
        if username == Config.LOGIN_USERNAME and password == Config.LOGIN_PASSWORD:
            session.permanent = True
            session['logged_in'] = True
            session['username'] = username
            return jsonify({'success': True, 'message': 'Login successful'}), 200
        else:
            return jsonify({'success': False, 'error': 'Invalid username or password'}), 401
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Logout route
@app.route('/logout')
def logout():
    """Handle logout"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    """Main page - shows navigation to 3 levels"""
    return render_template('index.html', states=INDIAN_STATES)

@app.route('/level1')
@login_required
def level1():
    """Level 1: Company Search (Location Search)"""
    # Cache-bust static assets so browser always loads latest JS/CSS after updates
    try:
        js_path = os.path.join(app.root_path, 'static', 'js', 'main.js')
        css_path = os.path.join(app.root_path, 'static', 'css', 'style.css')
        static_version = str(int(max(os.path.getmtime(js_path), os.path.getmtime(css_path))))
    except Exception:
        static_version = str(int(time.time()))

    return render_template('level1.html', static_version=static_version)

@app.route('/level2')
@login_required
def level2():
    """Level 2: Contact Enrichment (Contact Database)"""
    return render_template('level2.html')

@app.route('/api/level1/search', methods=['POST'])
def level1_search():
    """Level 1: Search companies using location search, save to database"""
    try:
        data = request.json
        project_name = data.get('project_name', '').strip()
        search_type = data.get('search_type', 'pin').strip()  # 'pin' or 'place'
        pin_codes_input = data.get('pin_code', '').strip()
        place_names_input = data.get('place_name', '').strip()
        industry = data.get('industry', '').strip()
        
        # Validate project name
        if not project_name:
            return jsonify({'error': 'Project name is required'}), 400
        
        # Validate project name format (alphanumeric, spaces, hyphens, underscores, forward slashes, parentheses)
        if not re.match(r'^[a-zA-Z0-9\s\-_/()]+$', project_name):
            invalid_chars = set(re.findall(r'[^a-zA-Z0-9\s\-_/()]', project_name))
            if invalid_chars:
                invalid_chars_str = ', '.join(f"'{c}'" for c in sorted(invalid_chars)[:5])
                return jsonify({
                    'error': f'Project name contains invalid characters: {invalid_chars_str}. Please use only letters, numbers, spaces, hyphens (-), underscores (_), forward slashes (/), and parentheses ().'
                }), 400
            return jsonify({'error': 'Project name can only contain letters, numbers, spaces, hyphens, underscores, forward slashes, and parentheses'}), 400
        
        if len(project_name) < 3:
            return jsonify({'error': 'Project name must be at least 3 characters'}), 400
        
        if len(project_name) > 100:
            return jsonify({'error': 'Project name must be less than 100 characters'}), 400
        
        # Handle PIN code search
        if search_type == 'pin':
            # Single PIN code only
            pin = pin_codes_input.strip()
            
            # Validate single 6-digit PIN
            if not pin or not pin.isdigit() or len(pin) != 6:
                return jsonify({'error': 'Please enter a valid 6-digit PIN code (e.g., 400001)'}), 400
            
            pin_codes = [pin]
        else:
            # Handle Place Name search
            place_names = [p.strip() for p in place_names_input.split(',') if p.strip()]
            if not place_names:
                return jsonify({'error': 'Please enter at least one place name (e.g., Mumbai, Delhi, Bangalore)'}), 400
            pin_codes = []  # Not used for place search
        
        # Safely get max_companies with error handling
        try:
            max_companies = int(data.get('max_companies', 20))
        except (ValueError, TypeError):
            max_companies = 20  # Default to 20 if conversion fails
        
        # Validate max_companies (limit to reasonable range)
        if max_companies < 1 or max_companies > 100:
            max_companies = 20  # Default to 20 if invalid
        
        logger.info(f"Level1 search: project={project_name}, type={search_type}, max={max_companies}")
        
        def generate():
            try:
                # Use project_name as the session identifier
                session_key = project_name
                
                # Initialize progress in Supabase
                initial_progress = {
                    'stage': 'searching_places',
                    'message': 'Searching locations...',
                    'current': 0,
                    'total': 0,
                    'companies_found': 0,
                    'status': 'in_progress'
                }
                try:
                    get_supabase_client().save_progress(session_key, initial_progress)
                except Exception as e:
                    logger.warning(f"Could not save progress to Supabase: {e}")
                
                yield f"data: {json.dumps({'type': 'progress', 'data': initial_progress})}\n\n"
                
                # Load companies already in DB for this project — don't show or save them again (same search = no duplicates)
                existing_in_db = get_supabase_client().get_level1_companies(project_name=project_name, selected_only=False, limit=5000)
                existing_place_ids = {c.get('place_id') for c in existing_in_db if c.get('place_id')}
                existing_fingerprints = {_company_fingerprint(c) for c in existing_in_db}
                if existing_in_db:
                    logger.info(f"Filtering out {len(existing_in_db)} existing companies for project '{project_name}'.")
                
                # Step 1: Search locations based on search type
                all_companies = []
                search_errors = []  # Track errors for better user feedback
                # Track seen IDs and fingerprints to prevent duplicates during progressive loading
                seen_place_ids_progressive = set()
                seen_fingerprints_progressive = set()
                
                if search_type == 'pin':
                    # === SIMPLE SEQUENTIAL SEARCH FOR PINS ===
                    # Search one PIN at a time until target reached or all PINs exhausted
                    total_locations = len(pin_codes)
                    logger.info(f"Sequential PIN search: {total_locations} PINs, target={max_companies}")
                    
                    for idx, pin_code in enumerate(pin_codes, 1):
                        if len(all_companies) >= max_companies:
                            break
                        
                        remaining_target = max_companies - len(all_companies)
                        logger.info(f"PIN {pin_code}: need {remaining_target} more companies")
                        
                        yield f"data: {json.dumps({'type': 'progress', 'data': {'stage': 'searching_places', 'message': f'Searching PIN {idx}/{total_locations}: {pin_code}...', 'current': idx, 'total': total_locations, 'companies_found': len(all_companies)}})}\n\n"
                        
                        try:
                            # Search this PIN until target reached or exhausted
                            companies_from_pin = 0
                            for company in search_pins_progressively(
                                pin_code=pin_code,
                                industry=industry,
                                max_results=remaining_target,
                                pin_idx=idx,
                                total_pins=total_locations
                            ):
                                # Deduplication checks
                                place_id = company.get('place_id')
                                if place_id and place_id in existing_place_ids:
                                    continue
                                if place_id and place_id in seen_place_ids_progressive:
                                    continue
                                if place_id:
                                    seen_place_ids_progressive.add(place_id)
                                
                                fp = _company_fingerprint(company)
                                if fp in existing_fingerprints:
                                    continue
                                if fp in seen_fingerprints_progressive:
                                    continue
                                if _is_same_company_by_name_address(company, all_companies, set()):
                                    continue
                                seen_fingerprints_progressive.add(fp)
                                
                                all_companies.append(company)
                                companies_from_pin += 1
                                
                                # Send to frontend
                                yield f"data: {json.dumps({'type': 'company_update', 'data': company, 'progress': {'current': len(all_companies), 'total': max_companies, 'companies_found': len(all_companies)}})}\n\n"
                                
                                if len(all_companies) >= max_companies:
                                    break
                            
                            logger.info(f"PIN {pin_code}: found {companies_from_pin} companies")
                            
                        except Exception as e:
                            logger.warning(f"Error searching PIN {pin_code}: {e}")
                            search_errors.append(f"PIN {pin_code}: {str(e)}")
                            continue
                    
                    logger.info(f"PIN search complete: {len(all_companies)} companies from {total_locations} PINs")
                else:
                    # === SIMPLE SEQUENTIAL SEARCH FOR PLACES ===
                    # Search one place at a time until target reached or all places exhausted
                    total_locations = len(place_names)
                    logger.info(f"Sequential Place search: {total_locations} places, target={max_companies}")
                    
                    for idx, place_name in enumerate(place_names, 1):
                        if len(all_companies) >= max_companies:
                            break
                        
                        remaining_target = max_companies - len(all_companies)
                        logger.info(f"Place {place_name}: need {remaining_target} more companies")
                        
                        yield f"data: {json.dumps({'type': 'progress', 'data': {'stage': 'searching_places', 'message': f'Searching Place {idx}/{total_locations}: {place_name}...', 'current': idx, 'total': total_locations, 'companies_found': len(all_companies)}})}\n\n"
                        
                        try:
                            # Search this place until target reached or exhausted
                            companies_from_place = 0
                            for company in search_places_progressively(
                                place_name=place_name,
                                industry=industry,
                                max_results=remaining_target,
                                place_idx=idx,
                                total_places=total_locations
                            ):
                                # Deduplication checks
                                place_id = company.get('place_id')
                                if place_id and place_id in existing_place_ids:
                                    continue
                                if place_id and place_id in seen_place_ids_progressive:
                                    continue
                                if place_id:
                                    seen_place_ids_progressive.add(place_id)
                                
                                fp = _company_fingerprint(company)
                                if fp in existing_fingerprints:
                                    continue
                                if fp in seen_fingerprints_progressive:
                                    continue
                                if _is_same_company_by_name_address(company, all_companies, set()):
                                    continue
                                seen_fingerprints_progressive.add(fp)
                                
                                all_companies.append(company)
                                companies_from_place += 1
                                
                                # Send to frontend
                                yield f"data: {json.dumps({'type': 'company_update', 'data': company, 'progress': {'current': len(all_companies), 'total': max_companies, 'companies_found': len(all_companies)}})}\n\n"
                                
                                if len(all_companies) >= max_companies:
                                    break
                            
                            logger.info(f"Place {place_name}: found {companies_from_place} companies")
                            
                        except Exception as e:
                            logger.warning(f"Error searching Place {place_name}: {e}")
                            search_errors.append(f"Place {place_name}: {str(e)}")
                            continue
                    
                    logger.info(f"Place search complete: {len(all_companies)} companies from {total_locations} places")
                
                # Hard cap: never pass more than max_companies (user selected 50 = show/save at most 50)
                all_companies = all_companies[:max_companies]
                
                # Final dedupe pass (safety for any that slipped through)
                seen_place_ids = set()
                seen_fingerprints = set()
                deduplicated_companies = []
                for company in all_companies:
                    place_id = company.get('place_id')
                    if place_id:
                        if place_id in seen_place_ids:
                            continue
                        seen_place_ids.add(place_id)
                    else:
                        company_key = f"{company.get('company_name', '')}_{company.get('address', '')}"
                        if company_key in seen_place_ids:
                            continue
                        seen_place_ids.add(company_key)
                    fp = _company_fingerprint(company)
                    if fp in seen_fingerprints or _is_same_company_by_name_address(company, deduplicated_companies, set()):
                        continue
                    seen_fingerprints.add(fp)
                    deduplicated_companies.append(company)
                
                companies = deduplicated_companies[:max_companies]
                companies_count = len(companies) if companies else 0
                logger.info(f"Level1 search completed: {companies_count} companies for project '{project_name}'")
                
                if not companies or companies_count == 0:
                    if search_type == 'pin':
                        location_str = ', '.join(pin_codes)
                        location_type_str = 'PIN code(s)'
                    else:
                        location_str = ', '.join(place_names)
                        location_type_str = 'Place(s)'
                    
                    # Provide better message based on what happened
                    if search_errors:
                        error_details = '; '.join(search_errors)
                        error_msg = f'No companies found for {location_type_str}: {location_str}. Errors: {error_details}. This may be due to service quota limits or network issues. Please try again in a few minutes.'
                    elif existing_in_db:
                        error_msg = f'No new companies — all results for these {location_type_str.lower()} are already in this campaign. Try different PINs, places, or industry.'
                    else:
                        error_msg = f'No companies found for {location_type_str}: {location_str}. Please try different locations or check if they are correct. You may also want to try a broader industry term.'
                    
                    logger.warning(f"No companies found for project '{project_name}': {location_str}. Errors: {search_errors}")
                    yield f"data: {json.dumps({'type': 'complete', 'data': {'companies': [], 'message': error_msg, 'total_companies': 0, 'errors': search_errors}})}\n\n"
                    return
                
                # Log company details for debugging
                logger.info(f"📋 Companies to save: {companies_count}")
                if companies:
                    logger.info(f"📋 First company sample: {companies[0].get('company_name', 'Unknown')} (place_id: {companies[0].get('place_id', 'None')})")
                
                # Update progress in Supabase
                saving_progress = {
                    'stage': 'saving',
                    'message': f'Found {len(companies)} companies. Saving to database...',
                    'current': 0,
                    'total': len(companies),
                    'companies_found': len(companies),
                    'status': 'in_progress'
                }
                try:
                    get_supabase_client().save_progress(session_key, saving_progress)
                except Exception as e:
                    logger.warning(f"Could not save progress to Supabase: {e}")
                yield f"data: {json.dumps({'type': 'progress', 'data': saving_progress})}\n\n"
                
                # Save to Supabase database
                try:
                    search_params = {
                        'project_name': project_name,  # User-defined project name
                        'industry': industry,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    # Store search parameters based on search type
                    if search_type == 'pin':
                        search_params['pin_codes'] = ','.join(pin_codes)  # Store all PIN codes as comma-separated
                    else:
                        search_params['place_names'] = ','.join(place_names)  # Store all place names as comma-separated
                    
                    logger.info(f"💾 Starting save operation for project '{project_name}' with {len(companies)} companies")
                    save_result = get_supabase_client().save_level1_results(companies, search_params)
                    
                    logger.info(f"💾 Save result: {save_result}")
                    
                    if save_result.get('success'):
                        saved_count = save_result.get('count', 0)
                        logger.info(f"Saved {saved_count} companies to Supabase for project '{project_name}'")
                        
                        # Double-check: if count is 0, that's a problem
                        if saved_count == 0:
                            error_msg = f"No companies were saved to database for project '{project_name}'. Save result: {save_result}"
                            logger.error(error_msg)
                            
                            # Try to verify what's in the database
                            try:
                                verify = get_supabase_client().client.table('level1_companies').select('id', count='exact').eq('project_name', project_name).execute()
                                db_count = verify.count if hasattr(verify, 'count') else (len(verify.data) if verify.data else 0)
                                logger.error(f"❌ Database verification: Found {db_count} companies for project '{project_name}'")
                            except Exception as verify_err:
                                logger.error(f"❌ Could not verify database: {verify_err}")
                            
                            raise Exception(error_msg)
                    else:
                        error_msg = save_result.get('error', 'Unknown error')
                        logger.error(f"Error saving to Supabase for project '{project_name}': {error_msg}")
                        raise Exception(f"Failed to save to Supabase: {error_msg}")
                except Exception as e:
                    error_msg = str(e)
                    logger.error(f"Save failed during search: {error_msg}")
                    import traceback
                    traceback.print_exc()
                    
                    # Send error to frontend so user knows save failed
                    yield f"data: {json.dumps({'type': 'error', 'data': {'error': f'Failed to save companies to database: {error_msg}. Companies were found but could not be saved.'}})}\n\n"
                    
                    # Still send companies so user can see them, but mark as not saved
                    yield f"data: {json.dumps({'type': 'complete', 'data': {'companies': companies, 'message': f'Found {len(companies)} companies but SAVE FAILED: {error_msg}. Please try saving again.', 'total_companies': len(companies), 'save_failed': True}})}\n\n"
                    return
                
                # Companies were already sent progressively during search (lazy loading)
                # No need to send them again - just update progress
                try:
                    # Update progress in Supabase
                    update_progress = {
                        'current': len(companies),
                        'message': f'Found {len(companies)} companies and saving to database...',
                        'status': 'in_progress'
                    }
                    get_supabase_client().save_progress(session_key, update_progress)
                except Exception as progress_err:
                    logger.warning(f"⚠️  Could not update progress: {progress_err}")
                
                # Final result - mark as completed in Supabase
                completed_progress = {
                    'stage': 'completed',
                    'message': f'Found {len(companies)} companies and saved to database.',
                    'current': len(companies),
                    'total': len(companies),
                    'companies_found': len(companies),
                    'status': 'completed'
                }
                
                # Verify save was successful before marking as completed
                try:
                    verify = get_supabase_client().client.table('level1_companies').select('id', count='exact').eq('project_name', project_name).execute()
                    actual_count = verify.count if hasattr(verify, 'count') else (len(verify.data) if verify.data else 0)
                    if actual_count == 0:
                        logger.error(f"❌ CRITICAL: Save reported success but database is empty for '{project_name}'")
                        yield f"data: {json.dumps({'type': 'error', 'data': {'error': 'Companies were not saved to database. Please try saving again.'}})}\n\n"
                        yield f"data: {json.dumps({'type': 'complete', 'data': {'companies': companies, 'message': f'Found {len(companies)} companies but SAVE FAILED. Please click Save button to retry.', 'total_companies': len(companies), 'save_failed': True}})}\n\n"
                        return
                except Exception as verify_err:
                    logger.warning(f"⚠️  Could not verify save: {verify_err}")
                get_supabase_client().save_progress(session_key, completed_progress)
                
                # When we got fewer than requested, say so so the client knows it's a limit of data, not a bug
                n = len(companies)
                if n < max_companies:
                    msg = f'Found {n} of up to {max_companies} requested (no more results for these locations). Saved to database. Proceed to Level 2 for contact enrichment.'
                else:
                    msg = f'Found {n} companies and saved to database. Proceed to Level 2 for contact enrichment.'
                result = {
                    'companies': companies,
                    'total_companies': n,
                    'max_requested': max_companies,
                    'message': msg
                }
                
                yield f"data: {json.dumps({'type': 'complete', 'data': result})}\n\n"
                
            except (BrokenPipeError, ConnectionResetError, GeneratorExit):
                # Client closed the connection (common with streaming responses). Not an app error.
                logger.info(f"Level 1: Client disconnected during stream for project '{project_name}'")
                return
            except Exception as e:
                error_msg = str(e)
                logger.error(f"❌ Error in Level 1 search stream: {error_msg}")
                import traceback
                traceback.print_exc()
                
                # Clean error message - remove Python internals
                clean_error = error_msg
                if 'logger' in clean_error.lower() or 'NameError' in clean_error:
                    clean_error = "An internal error occurred. Please try again or contact support."
                if 'Traceback' in clean_error or 'File "' in clean_error:
                    clean_error = "An error occurred during processing. Please try again."
                
                try:
                    yield f"data: {json.dumps({'type': 'error', 'data': {'error': clean_error}})}\n\n"
                    # Send a final complete message to close the stream
                    yield f"data: {json.dumps({'type': 'complete', 'data': {'companies': [], 'message': clean_error, 'total_companies': 0, 'save_failed': True}})}\n\n"
                except (BrokenPipeError, ConnectionResetError, GeneratorExit):
                    return
                except Exception as send_err:
                    logger.error(f"❌ Could not send error to client: {send_err}")
                    return
            finally:
                # Always ensure stream is properly closed
                try:
                    # Send a final message to ensure stream completes
                    # This prevents ERR_INCOMPLETE_CHUNKED_ENCODING
                    pass
                except:
                    pass
                # Clean up progress from Supabase after a delay (keep for 1 hour for recovery)
                # For immediate cleanup, uncomment the line below:
                # get_supabase_client().delete_progress(session_key)
        
        # Wrap generator to ensure it always completes
        def safe_generate():
            try:
                for chunk in generate():
                    yield chunk
            except (BrokenPipeError, ConnectionResetError, GeneratorExit):
                # Client disconnected - this is normal
                pass
            except Exception as e:
                logger.error(f"❌ Generator error: {e}")
                # Send final error message
                try:
                    yield f"data: {json.dumps({'type': 'error', 'data': {'error': 'An error occurred during streaming'}})}\n\n"
                except:
                    pass
        
        return Response(stream_with_context(safe_generate()), mimetype='text/event-stream', headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'  # Disable buffering for nginx
        })
        
    except Exception as e:
        logger.error(f"Level 1 search error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/search/sync', methods=['POST'])
def search_sync():
    """Legacy synchronous endpoint (kept for compatibility)"""
    try:
        data = request.json
        pin_codes_input = data.get('pin_code', '').strip()
        industry = data.get('industry', '').strip()
        
        # Parse multiple PIN codes (comma-separated) and auto-complete incomplete ones
        all_pins = [p.strip() for p in pin_codes_input.split(',') if p.strip()]
        
        # Separate into valid (6 digits), incomplete (numeric but < 6 digits), and invalid
        valid_pins = []
        incomplete_pins = []
        invalid_pins = []
        
        for pin in all_pins:
            if pin.isdigit() and len(pin) == 6:
                valid_pins.append(pin)
            elif pin.isdigit() and len(pin) < 6:
                incomplete_pins.append(pin)
            else:
                invalid_pins.append(pin)
        
        # Auto-complete incomplete PIN codes using prefix from first valid PIN
        if valid_pins and incomplete_pins:
            first_valid = valid_pins[0]
            for incomplete in incomplete_pins:
                # Calculate how many digits we need from the prefix
                digits_needed = 6 - len(incomplete)
                if digits_needed > 0 and digits_needed <= 6:
                    prefix = first_valid[:digits_needed]
                    completed = prefix + incomplete
                    if len(completed) == 6 and completed.isdigit():
                        valid_pins.append(completed)
                        logger.debug(f"PIN auto-completed '{incomplete}' to '{completed}'")
        
        # Final list of PIN codes to use
        pin_codes = valid_pins
        
        if not pin_codes:
            error_msg = 'No valid PIN codes found. '
            if incomplete_pins:
                error_msg += f'Incomplete: {", ".join(incomplete_pins)} (need at least one full 6-digit PIN to auto-complete). '
            if invalid_pins:
                error_msg += f'Invalid: {", ".join(invalid_pins)}. '
            error_msg += 'Please enter at least one valid 6-digit PIN code.'
            return jsonify({'error': error_msg}), 400
        
        # Search all PIN codes
        all_companies = []
        # Request more companies per PIN to account for duplicates (request 1.5x to ensure we get enough unique results)
        max_companies_sync = 20
        companies_per_pin = max(1, int((max_companies_sync * 1.5) // len(pin_codes)))  # Request extra to account for duplicates
        
        for pin_code in pin_codes:
            companies = google_client.search_by_pin_and_industry(
                pin_code=pin_code,
                industry=industry,
                max_results=companies_per_pin
            )
            all_companies.extend(companies)
        
        # Deduplicate companies by place_id (Google's unique identifier)
        # This prevents the same company from appearing multiple times when searching multiple PIN codes
        seen_place_ids = set()
        deduplicated_companies = []
        for company in all_companies:
            place_id = company.get('place_id')
            if place_id and place_id not in seen_place_ids:
                seen_place_ids.add(place_id)
                deduplicated_companies.append(company)
            elif not place_id:
                # If no place_id, use company_name + address as fallback identifier
                company_key = f"{company.get('company_name', '')}_{company.get('address', '')}"
                if company_key not in seen_place_ids:
                    seen_place_ids.add(company_key)
                    deduplicated_companies.append(company)
        
        logger.debug(f"Deduplication: {len(all_companies)} → {len(deduplicated_companies)} unique")
        
        companies = deduplicated_companies[:20]  # Limit to 20 total
        
        if not companies:
            return jsonify({
                'companies': [],
                'message': 'No companies found for the given location'
            }), 200
        
        enriched_companies = apollo_client.enrich_company_data(companies)
        
        result = {
            'companies': enriched_companies,
            'total_companies': len(enriched_companies),
            'total_contacts': sum(len(c.get('people', [])) for c in enriched_companies)
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        logger.error(f"Level 1 search error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/process', methods=['POST'])
def level2_process():
    """Level 2: Process companies from Supabase with contact database to get contacts"""
    try:
        data = request.json
        batch_size = int(data.get('batch_size', 10))  # Process 10 companies per batch
        batch_number = int(data.get('batch_number', 1))
        project_name = data.get('project_name')
        designation = data.get('designation', '').strip()  # Custom designation/titles
        
        employee_ranges = _normalize_employee_ranges(data)

        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400

        try:
            health_response = requests.get("https://api.apollo.io/v1/auth/health", headers=apollo_client.headers, timeout=5)
            if health_response.status_code != 200:
                return jsonify({'error': f'Database connection failed (status {health_response.status_code}). Check config.'}), 401
        except Exception as e:
            return jsonify({'error': f'Database connection error: {str(e)}'}), 500

        # Get all companies for this project first
        all_companies = get_supabase_client().get_level1_companies(
            project_name=project_name,
            selected_only=False,
            include_excluded=False,
            limit=500,
        )

        # If frontend sent selected IDs, filter to just those
        selected_place_ids = data.get('selected_place_ids', [])
        if selected_place_ids:
            id_set = set(str(pid) for pid in selected_place_ids if pid)
            companies = [
                c for c in all_companies
                if str(c.get('place_id', '')) in id_set
                or str(c.get('company_name', '')) in id_set
            ]
            logger.info(f"Level 2: {len(companies)} companies matched from {len(selected_place_ids)} inline ids")
        else:
            # No IDs sent (old frontend cache or no selection) — use all companies
            companies = all_companies
            logger.info(f"Level 2: using all {len(companies)} companies for project '{project_name}'")

        if not companies:
            return jsonify({'error': 'No companies selected for Level 2. Please select companies first.'}), 400
        
        if employee_ranges:
            # First, fetch employee counts for companies that don't have them yet
            # IMPORTANT: Only fetch if NOT in database (saves API credits!)
            companies_without_employee_data = [c for c in companies if not c.get('total_employees')]
            if companies_without_employee_data:
                logger.info(f"Level 2: Fetching employee counts for {len(companies_without_employee_data)} companies")
                fetched_count = 0
                for company in companies_without_employee_data:
                    company_name = company.get('company_name', '')
                    website = company.get('website', '')
                    if company_name:
                        try:
                            total_employees = apollo_client.get_company_total_employees(company_name, website) or ''
                            if total_employees:
                                # Validate employee count before using it
                                # Check if it's a reasonable number (not corrupted data)
                                try:
                                    # Try to parse and validate
                                    cleaned = str(total_employees).replace(',', '').replace(' ', '').strip()
                                    # If it's a simple number, check it's reasonable
                                    if cleaned.isdigit():
                                        num = int(cleaned)
                                        if num <= 0 or num > 1000000:
                                            total_employees = ''
                                except:
                                    pass  # If validation fails, still use the original value
                                
                                if total_employees:
                                    company['total_employees'] = total_employees
                                    fetched_count += 1
                                    # Update in database for future use
                                    if company.get('place_id'):
                                        try:
                                            get_supabase_client().update_level1_company_metrics(
                                                project_name=project_name,
                                                place_id=company['place_id'],
                                                total_employees=total_employees
                                            )
                                        except:
                                            pass  # Best effort - don't fail if update fails
                            else:
                                pass
                        except Exception:
                            pass
                logger.info(f"Level 2: Fetched employee data for {fetched_count}/{len(companies_without_employee_data)} companies")
            companies_before_filter_list = companies.copy()
            companies_before_filter_count = len(companies)
            companies = filter_companies_by_employee_range(companies, employee_ranges)
            companies_after_filter = len(companies)
            logger.info(f"Level 2: After employee filter: {companies_after_filter} companies")
            
            if not companies:
                # Check if any companies had employee data
                companies_with_data = [c for c in companies_before_filter_list if c.get('total_employees')]
                # Ensure employee_ranges is a list before joining (extra safety check)
                if not isinstance(employee_ranges, (list, tuple)):
                    employee_ranges = []
                ranges_str = ', '.join(str(r) for r in employee_ranges) if employee_ranges else ''
                if not companies_with_data:
                    error_msg = f'No companies have employee data available. Employee range filter(s) "{ranges_str}" require employee data, but none of the {companies_before_filter_count} companies have this information in the database. Please select "All Company Sizes" to process all companies regardless of employee count.'
                else:
                    error_msg = f'No companies found matching employee range(s): {ranges_str}. {companies_before_filter_count - companies_after_filter} companies were filtered out. Try selecting "All Company Sizes" or different employee ranges.'
                
                return jsonify({
                    'message': error_msg,
                    'completed': True,
                    'processed': 0,
                    'total_companies': 0,
                    'error': f'Employee range filter removed all companies'
                }), 200
        
        # Process all companies sequentially with real-time SSE updates (like Level 1)
        def generate():
            try:
                total_companies = len(companies)
                total_contacts = 0
                enriched_companies = []
                default_batch_name = f"{project_name}_Main_Batch"
                
                # Send initial progress
                yield f"data: {json.dumps({'type': 'progress', 'data': {'stage': 'processing', 'message': f'Starting to process {total_companies} companies...', 'current': 0, 'total': total_companies, 'contacts_found': 0}})}\n\n"
                
                # Process companies one by one with real-time updates
                for idx, company in enumerate(companies, 1):
                    company_name = company.get('company_name', '')
                    website = company.get('website', '')
                    place_id = company.get('place_id', '')
                    
                    # Send progress update
                    yield f"data: {json.dumps({'type': 'progress', 'data': {'stage': 'processing', 'message': f'Processing: {company_name}... ({idx}/{total_companies})', 'current': idx, 'total': total_companies, 'contacts_found': total_contacts}})}\n\n"

                    titles = [t.strip() for t in designation.split(',') if t.strip()] if (designation and designation.strip()) else None
                    existing_contacts = get_supabase_client().get_contacts_by_company(company_name, project_name, titles)

                    if existing_contacts:
                        people = existing_contacts
                        logger.debug(f"Level 2: {company_name} — using {len(people)} existing contacts")
                    else:
                        if website and website.strip():
                            try:
                                people = apollo_client.search_people_by_company(company_name, website, titles=titles)
                                if not people:
                                    people = apollo_client.search_people_by_company_name(company_name, titles=titles)
                            except Exception as e:
                                logger.warning(f"Level 2: Apollo search failed for {company_name}: {e}")
                                people = []
                        else:
                            try:
                                people = apollo_client.search_people_by_company_name(company_name, titles=titles)
                            except Exception as e:
                                logger.warning(f"Level 2: Apollo search failed for {company_name}: {e}")
                                people = []
                        if people:
                            logger.debug(f"Level 2: {company_name} — {len(people)} contacts")
                    
                    # Get employee count (use existing data first to save credits!)
                    # OPTIMIZATION: Only fetch if employee range filter is selected
                    total_employees = company.get('total_employees', '') or ''
                    
                    # Only fetch employee count if user selected employee range filter
                    # If no filter selected, skip API call to save credits
                    if employee_ranges and len(employee_ranges) > 0:
                        # User selected employee filter - we need employee count
                        if not total_employees:
                            total_employees = apollo_client.get_company_total_employees(company_name, website) or ''
                            if total_employees:
                                company['total_employees'] = total_employees
                                # Save to database immediately to avoid future API calls
                                if company.get('place_id'):
                                    try:
                                        get_supabase_client().update_level1_company_metrics(
                                            project_name=project_name,
                                            place_id=company['place_id'],
                                            total_employees=total_employees
                                        )
                                    except Exception:
                                        pass
                        else:
                            pass
                    else:
                        pass
                    active_members = len(people or [])
                    active_members_with_email = sum(1 for p in (people or []) if p.get('email'))
                    total_contacts += active_members
                    
                    enriched_company = {
                        'company_name': company_name,
                        'address': company.get('address', ''),
                        'website': website,
                        'phone': company.get('phone', ''),
                        'pin_code': company.get('pin_code', ''),
                        'industry': company.get('industry', ''),
                        'total_employees': total_employees,
                        'active_members': active_members,
                        'active_members_with_email': active_members_with_email,
                        'people': people,
                        'place_id': place_id,
                        'founders': [p for p in (people or []) if p.get('title') and any(keyword.lower() in p.get('title', '').lower() 
                                                       for keyword in ['founder', 'owner', 'ceo', 'co-founder', 'founder/owner'])],
                        'hr_contacts': [p for p in (people or []) if p.get('title') and any(keyword.lower() in p.get('title', '').lower() 
                                                      for keyword in ['hr', 'human resources', 'recruiter', 'talent', 'human resource'])]
                    }
                    
                    enriched_companies.append(enriched_company)
                    
                    # Update metrics in database
                    if place_id:
                        try:
                            get_supabase_client().update_level1_company_metrics(
                                project_name=project_name,
                                place_id=place_id,
                                total_employees=total_employees,
                                active_members=active_members,
                                active_members_with_email=active_members_with_email,
                            )
                        except:
                            pass
                    
                    # Send company update in real-time
                    yield f"data: {json.dumps({'type': 'company_update', 'data': enriched_company, 'progress': {'current': idx, 'total': total_companies, 'contacts_found': total_contacts}})}\n\n"
                
                # NOTE: Filtering already happens BEFORE enrichment in apollo_client.py
                # So contacts here are already filtered - no need to filter again (saves credits!)
                # Only do final safety check to exclude generic employee titles
                if designation and designation.strip():
                    user_titles = [t.strip().lower() for t in designation.split(',') if t.strip()]
                    excluded_titles = ['employee', 'staff', 'worker', 'member', 'personnel']
                    
                    # Final safety filter: Only exclude generic employee titles (contacts already filtered before enrichment)
                    filtered_total = 0
                    for company in enriched_companies:
                        original_count = len(company.get('people', []))
                        filtered_people = []
                        
                        for person in company.get('people', []):
                            person_title = (person.get('title', '') or '').lower().strip()
                            
                            # Only exclude generic employee titles (contacts already matched user's designation before enrichment)
                            if person_title and person_title not in excluded_titles:
                                filtered_people.append(person)
                        
                        company['people'] = filtered_people
                        filtered_total += len(filtered_people)
                        
                        if original_count != len(filtered_people):
                            logger.debug(f"Level 2: {company.get('company_name')} — filtered {original_count} → {len(filtered_people)}")
                    total_contacts = filtered_total
                else:
                    pass
                # Save filtered results to Supabase
                yield f"data: {json.dumps({'type': 'progress', 'data': {'stage': 'saving', 'message': 'Saving contacts to database...', 'current': total_companies, 'total': total_companies, 'contacts_found': total_contacts}})}\n\n"
                
                save_result = get_supabase_client().save_level2_results(
                    enriched_companies, 
                    project_name=project_name,
                    batch_name=default_batch_name
                )
                
                if not save_result.get('success'):
                    err = save_result.get('error', 'Unknown error saving to database')
                    logger.error(f"Level 2 save failed: {err}")
                    yield f"data: {json.dumps({'type': 'error', 'data': {'error': f'Could not save to database: {err}'}})}\n\n"
                    return
                
                logger.info(f"Level 2 complete: {total_companies} companies, {total_contacts} contacts saved to batch '{default_batch_name}'")
                # Send completion (only after save succeeded)
                yield f"data: {json.dumps({'type': 'complete', 'data': {'total_companies': total_companies, 'total_contacts': total_contacts, 'message': f'Found {total_contacts} contacts from {total_companies} companies', 'batch_name': default_batch_name}})}\n\n"
                
            except (BrokenPipeError, ConnectionResetError, GeneratorExit):
                logger.info("Level 2: Client disconnected during stream")
                return
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Level 2 stream error: {error_msg}", exc_info=True)
                try:
                    yield f"data: {json.dumps({'type': 'error', 'data': {'error': error_msg}})}\n\n"
                except:
                    pass
        
        return Response(stream_with_context(generate()), mimetype='text/event-stream', headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'
        })
        
    except Exception as e:
        logger.error(f"Level 2 processing error: {e}", exc_info=True)
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/level1/save', methods=['POST'])
def level1_save_manual():
    """Manual save endpoint - saves companies that were found but not saved"""
    try:
        data = request.json
        project_name = data.get('project_name', '').strip()
        companies = data.get('companies', [])
        pin_codes = data.get('pin_codes', '')
        industry = data.get('industry', '')
        
        if not project_name:
            return jsonify({'success': False, 'error': 'Project name is required'}), 400
        
        if not companies:
            return jsonify({'success': False, 'error': 'No companies to save'}), 400
        
        logger.info(f"💾 Manual save requested for project '{project_name}' with {len(companies)} companies")
        
        search_params = {
            'project_name': project_name,
            'pin_codes': pin_codes or '',
            'industry': industry or '',
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        save_result = get_supabase_client().save_level1_results(companies, search_params)
        
        if save_result.get('success'):
            saved_count = save_result.get('count', 0)
            logger.info(f"✅ Manual save successful: {saved_count} companies saved for project '{project_name}'")
            return jsonify({
                'success': True,
                'count': saved_count,
                'message': f'Successfully saved {saved_count} companies to database'
            }), 200
        else:
            error_msg = save_result.get('error', 'Unknown error')
            logger.error(f"❌ Manual save failed for project '{project_name}': {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 500
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"❌ Error in manual save: {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': error_msg}), 500

@app.route('/api/level1/projects', methods=['GET'])
def get_projects_list():
    """Get list of all projects (for history/resume feature) - from Supabase"""
    try:
        client = get_supabase_client()
        projects = client.get_projects_list()
        logger.debug(f"API: Returning {len(projects)} projects")
        return jsonify({'projects': projects}), 200
    except Exception as e:
        print(f"❌ Error getting projects list: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'projects': [], 'error': str(e)}), 200

@app.route('/api/level1/project-data', methods=['GET'])
def get_project_data():
    """Get project data including companies, PIN codes, and industry for a specific project"""
    try:
        project_name = request.args.get('project_name')
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        
        # Get project details from projects list
        projects = get_supabase_client().get_projects_list()
        project_info = next((p for p in projects if p.get('project_name') == project_name), None)
        
        if not project_info:
            return jsonify({'error': 'Project not found'}), 404
        
        # Get companies for this project
        companies = get_supabase_client().get_level1_companies(project_name=project_name, selected_only=False, limit=100)
        
        # Format companies for frontend
        formatted_companies = []
        for company in companies:
            formatted_companies.append({
                'company_name': company.get('company_name', ''),
                'website': company.get('website', ''),
                'phone': company.get('phone', ''),
                'address': company.get('address', ''),
                'pin_code': company.get('pin_code', ''),
                'industry': company.get('industry', ''),
                'place_type': company.get('place_type', ''),
                'place_id': company.get('place_id', ''),
                'business_status': company.get('business_status', ''),
                'selected_for_level2': company.get('selected_for_level2', False)
            })
        
        return jsonify({
            'project_name': project_name,
            'pin_codes': project_info.get('pin_codes', ''),
            'industry': project_info.get('industry', ''),
            'search_date': project_info.get('search_date', ''),
            'companies': formatted_companies,
            'total_companies': len(formatted_companies)
        }), 200
        
    except Exception as e:
        print(f"Error getting project data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/level1/excluded-companies', methods=['GET'])
def get_level1_excluded_companies():
    """Get companies that were soft-deleted (excluded) for a project - for Level 1 View Deleted."""
    try:
        project_name = request.args.get('project_name')
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        limit = int(request.args.get('limit', 500))
        companies = get_supabase_client().get_level1_companies(
            project_name=project_name, excluded_only=True, limit=limit
        )
        formatted = []
        for c in companies:
            formatted.append({
                'company_name': c.get('company_name', ''),
                'website': c.get('website', ''),
                'phone': c.get('phone', ''),
                'address': c.get('address', ''),
                'industry': c.get('industry', ''),
                'place_id': c.get('place_id', ''),
            })
        return jsonify({'companies': formatted}), 200
    except Exception as e:
        logger.error(f"Error fetching excluded companies: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/level1/delete-companies', methods=['POST'])
def delete_level1_companies():
    """Delete selected Level 1 companies from Supabase (per project)."""
    try:
        data = request.json or {}
        project_name = (data.get('project_name') or '').strip()
        identifiers = data.get('identifiers', [])

        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        if not isinstance(identifiers, list) or len(identifiers) == 0:
            return jsonify({'error': 'identifiers must be a non-empty list'}), 400

        result = get_supabase_client().delete_level1_companies(project_name=project_name, identifiers=identifiers)
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Failed to delete companies')}), 500

        return jsonify(result), 200

    except Exception as e:
        print(f"Error deleting Level 1 companies: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/level1/delete-project', methods=['POST'])
def delete_level1_project():
    """Delete an entire project from Supabase (Level 1 companies + Level 2 contacts)."""
    try:
        data = request.json or {}
        project_name = (data.get('project_name') or '').strip()

        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400

        result = get_supabase_client().delete_project(project_name=project_name)
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Failed to delete project')}), 500

        return jsonify(result), 200

    except Exception as e:
        print(f"Error deleting project: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/level1/check-project', methods=['POST'])
def check_project_exists():
    """Check if project name already exists"""
    try:
        data = request.json
        project_name = data.get('project_name', '').strip()
        
        if not project_name:
            return jsonify({'exists': False}), 200
        
        # Check in Supabase
        companies = get_supabase_client().get_level1_companies(project_name=project_name, limit=1)
        exists = len(companies) > 0
        
        return jsonify({'exists': exists}), 200
    except Exception as e:
        print(f"Error checking project: {str(e)}")
        return jsonify({'exists': False}), 200

@app.route('/api/level1/select-for-level2', methods=['POST'])
def select_companies_for_level2():
    """Mark selected companies for Level 2 processing"""
    try:
        data = request.json
        companies = data.get('companies', [])
        project_name = data.get('project_name', '').strip()
        
        if not companies:
            return jsonify({'error': 'No companies provided'}), 400
        
        if not project_name:
            return jsonify({'error': 'Project name is required'}), 400
        
        # Mark companies as selected in Supabase (with project_name for filtering)
        result = get_supabase_client().mark_companies_selected(companies, project_name=project_name)
        
        if result.get('success'):
            return jsonify({
                'status': 'success',
                'message': f'Successfully marked {len(companies)} companies for Level 2',
                'count': len(companies),
                'project_name': project_name
            }), 200
        else:
            return jsonify({'error': result.get('error', 'Failed to mark companies')}), 500
            
    except Exception as e:
        print(f"Error marking companies for Level 2: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/companies', methods=['GET'])
def get_level2_companies():
    """Level 2: Get companies (selected + not excluded), excluded list, and no-database-data list.
    When batch_name is provided, 'Not Found in database' is scoped to the current batch only
    (companies in project that have no contacts in this batch); old data does not persist."""
    try:
        project_name = request.args.get('project_name')
        batch_name = request.args.get('batch_name')
        limit = int(request.args.get('limit', 500))

        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400

        client = get_supabase_client()
        # Main list: try "selected" first (respect explicit selection if present)
        # IMPORTANT: Do NOT include soft-excluded companies here.
        main_companies = client.get_level1_companies(
            project_name=project_name,
            selected_only=True,
            include_excluded=False,
            limit=limit,
        )

        # Fallback: if nothing has been explicitly selected yet,
        # load all non-excluded companies for this project so the user
        # can still proceed without having to re-select in Level 1.
        if not main_companies:
            main_companies = client.get_level1_companies(
                project_name=project_name,
                selected_only=False,
                include_excluded=False,
                limit=limit,
            )

        # Excluded (soft-removed) companies
        excluded_companies = client.get_level1_companies(
            project_name=project_name,
            excluded_only=True,
            limit=limit,
        )

        # IMPORTANT: Never auto-promote excluded companies into the main list.
        # Not Found in database: when batch_name is set, show only companies from current batch
        # that have no contacts in this batch (so list refreshes on re-run and is batch-specific).
        if batch_name:
            companies_with_contacts_in_batch = set(
                c.lower() for c in client.get_company_names_with_contacts_in_batch(batch_name)
            )
            no_apollo_data_companies = [
                c for c in main_companies
                if (c.get('company_name') or '').strip().lower() not in companies_with_contacts_in_batch
            ]
        else:
            # No batch: use project-level (database lookup returned 0 contacts)
            no_apollo_data_companies = client.get_level1_companies(
                project_name=project_name,
                apollo_no_data_only=True,
                limit=limit,
            )

        def to_ui(c, selected=True, default_selected=True):
            place_id = c.get('place_id') or c.get('company_name')
            return {
                'place_id': place_id,
                'company_name': c.get('company_name', ''),
                'website': c.get('website', ''),
                'phone': c.get('phone', ''),
                'address': c.get('address', ''),
                'pin_code': c.get('pin_code', ''),
                'industry': c.get('industry', ''),
                'total_employees': c.get('total_employees', ''),
                'active_members': c.get('active_members', 0),
                'active_members_with_email': c.get('active_members_with_email', 0),
                'selected': selected,
                'default_selected': default_selected,
            }

        companies = [to_ui(c) for c in main_companies]
        excluded_ui = [to_ui(c, selected=False, default_selected=False) for c in excluded_companies]
        no_apollo_ui = [to_ui(c, selected=False, default_selected=False) for c in no_apollo_data_companies]

        # Mode flag for frontend: whether we are showing explicit selection
        # or a fallback "all companies" list.
        mode = 'selected' if any(c.get('selected_for_level2') for c in main_companies) else 'all'
        return jsonify({
            'companies': companies,
            'excluded_companies': excluded_ui,
            'no_apollo_data_companies': no_apollo_ui,
            'project_name': project_name,
            'mode': mode,
            'total_companies_all': len(main_companies) + len(excluded_companies),
            'total_companies_selected': len(main_companies),
            'total_excluded': len(excluded_companies),
            'total_no_apollo_data': len(no_apollo_data_companies),
        }), 200
    except Exception as e:
        print(f"Error getting Level 2 companies: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/selection', methods=['POST'])
def set_level2_selection():
    """Level 2: Persist the user's selection (after removing companies) to Supabase"""
    try:
        data = request.json or {}
        project_name = data.get('project_name')
        selected_place_ids = data.get('selected_place_ids', [])

        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400

        if not isinstance(selected_place_ids, list):
            return jsonify({'error': 'selected_place_ids must be a list'}), 400

        # Persist selection (sets selected_for_level2 true/false for this project)
        result = get_supabase_client().set_level2_selection(project_name=project_name, selected_place_ids=selected_place_ids)
        
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Failed to save selection')}), 500

        return jsonify(result), 200
    except Exception as e:
        print(f"Error setting Level 2 selection: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/history', methods=['GET'])
def get_search_history():
    """Get search history from Supabase (same as projects list)"""
    try:
        projects = get_supabase_client().get_projects_list()
        # Convert to history format for compatibility
        history = [{
            'project_name': p.get('project_name', ''),
            'search_date': p.get('search_date', ''),
            'industry': p.get('industry', ''),
            'pin_codes': p.get('pin_codes', '')
        } for p in projects]
        return jsonify({'history': history}), 200
    except Exception as e:
        print(f"Error getting search history: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/status', methods=['GET'])
def level2_status():
    """Get status of companies in Supabase for Level 2"""
    try:
        project_name = request.args.get('project_name')
        
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        
        all_companies = get_supabase_client().get_level1_companies(project_name=project_name, selected_only=False, limit=50)
        selected = [c for c in all_companies if c.get('selected_for_level2', False)]
        selected_count = len(selected)
        total_count = len(all_companies)
        active_count = selected_count if selected_count > 0 else total_count

        # Attach project metadata so UI can confirm we are not mixing Industry/PINs
        session_meta = {}
        try:
            projects = get_supabase_client().get_projects_list()
            for p in projects:
                if p.get('project_name') == project_name:
                    session_meta = {
                        'pin_codes': p.get('pin_codes', ''),
                        'industry': p.get('industry', ''),
                        'search_date': p.get('search_date', ''),
                    }
                    break
        except Exception:
            session_meta = {}

        return jsonify({
            'total_companies': active_count,
            'total_companies_all': total_count,
            'total_companies_selected': selected_count,
            'mode': 'selected' if selected_count > 0 else 'all',
            'project_name': project_name,
            'session': session_meta,
            'ready_for_processing': active_count > 0
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/contacts', methods=['GET'])
def level2_contacts():
    """Get Level 2 contacts for a project"""
    try:
        project_name = request.args.get('project_name')
        batch_name = request.args.get('batch_name')  # Optional: filter by specific batch
        designation = request.args.get('designation', '').strip()  # User's designation filter
        
        if not project_name and not batch_name:
            return jsonify({'error': 'project_name or batch_name is required'}), 400
        
        # Log the designation being used
        # Note: For new batches, contacts are already filtered at save time, but we still filter here
        # for backward compatibility with existing batches that may have unfiltered data
        if designation:
            logger.debug(f"Level 2 contacts API: filtering by designation")
        else:
            pass
        
        # Get contacts from Supabase with designation filter
        # (Filtering here ensures backward compatibility with old batches, but new batches are already filtered)
        if batch_name:
            contacts = get_supabase_client().get_contacts_for_level3(batch_name=batch_name, designation=designation if designation else None)
        else:
            contacts = get_supabase_client().get_contacts_for_level3(project_name=project_name, designation=designation if designation else None)
        
        logger.debug(f"Level 2 contacts API: returning {len(contacts)} contacts")
        
        # Map phone_number to phone for frontend compatibility
        formatted_contacts = []
        for contact in contacts:
            formatted_contact = dict(contact)
            # Ensure phone field is available (map from phone_number if needed)
            phone_value = formatted_contact.get('phone') or formatted_contact.get('phone_number') or ''
            formatted_contact['phone'] = phone_value
            formatted_contact['phone_number'] = phone_value  # Keep both for compatibility
            # Also ensure contact_name maps to name
            if not formatted_contact.get('name') and formatted_contact.get('contact_name'):
                formatted_contact['name'] = formatted_contact['contact_name']
            formatted_contacts.append(formatted_contact)
        
        return jsonify({
            'success': True,
            'contacts': formatted_contacts,
            'count': len(formatted_contacts)
        }), 200
    except Exception as e:
        print(f"Error getting Level 2 contacts: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/enrich-phones', methods=['POST'])
def enrich_phones_parallel():
    """
    DEPRECATED: Phone numbers are no longer enriched via API to save credits.
    Phone numbers should be revealed in database dashboard when needed.
    This endpoint returns empty results with a message.
    """
    # Phone numbers are not requested via API to save credits
    # Users should reveal phone numbers in the outreach dashboard when needed
    return jsonify({
        'success': True,
        'phones': {},
        'message': 'Phone numbers are not enriched via API to save credits. Please reveal phone numbers in the database dashboard when needed.',
        'note': 'This saves ~2-3 credits per contact. Reveal phone numbers in the database dashboard for contacts you want to call.'
    }), 200

@app.route('/api/level2/save-batch', methods=['POST'])
def level2_save_batch():
    """Save Level 2 batch with batch_name"""
    try:
        data = request.json or {}
        project_name = data.get('project_name')
        batch_name = data.get('batch_name')
        
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        if not batch_name:
            return jsonify({'error': 'batch_name is required'}), 400
        
        # Update all contacts for this project to use the batch_name
        # This is a simple implementation - in production you might want more sophisticated logic
        result = get_supabase_client().update_batch_name(project_name, batch_name)
        
        return jsonify({
            'success': True,
            'message': f'Batch "{batch_name}" saved successfully',
            'batch_name': batch_name
        }), 200
    except Exception as e:
        print(f"Error saving batch: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/rename-batch', methods=['POST'])
def level2_rename_batch():
    """Rename a batch (update all contacts from old name to new name)"""
    try:
        data = request.json or {}
        old_name = (data.get('old_batch_name') or '').strip()
        new_name = (data.get('new_batch_name') or '').strip()
        if not old_name or not new_name:
            return jsonify({'error': 'old_batch_name and new_batch_name are required'}), 400
        result = get_supabase_client().rename_batch(old_name, new_name)
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Rename failed')}), 500
        return jsonify({
            'success': True,
            'message': f'Batch renamed to "{new_name}"',
            'batch_name': result.get('batch_name', new_name),
            'count': result.get('count', 0)
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/batches', methods=['GET'])
def level2_batches():
    """Get list of all Level 2 batches"""
    try:
        project_name = request.args.get('project_name')  # Optional: filter by project
        
        # Get batches from Supabase
        batches = get_supabase_client().get_batches_list(project_name=project_name)
        
        return jsonify({
            'success': True,
            'batches': batches,
            'count': len(batches)
        }), 200
    except Exception as e:
        print(f"Error getting batches: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/delete-contact', methods=['POST'])
def level2_delete_contact():
    """Soft-delete a contact: set deleted_at so it stays in DB and shows in View Deleted Contacts."""
    try:
        data = request.json or {}
        contact_id = data.get('contact_id')
        if contact_id is None:
            return jsonify({'error': 'contact_id is required'}), 400
        result = get_supabase_client().delete_level2_contact(contact_id)
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Failed to remove contact')}), 500
        return jsonify({'success': True, 'message': 'Contact moved to deleted list'}), 200
    except Exception as e:
        logger.error(f"Error soft-deleting contact: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/deleted-contacts', methods=['GET'])
def level2_deleted_contacts():
    """Get soft-deleted contacts for a batch or project (for View Deleted Contacts)."""
    try:
        batch_name = request.args.get('batch_name')
        project_name = request.args.get('project_name')
        if not batch_name and not project_name:
            return jsonify({'error': 'batch_name or project_name is required'}), 400
        contacts = get_supabase_client().get_deleted_level2_contacts(batch_name=batch_name, project_name=project_name)
        for c in contacts:
            c.setdefault('phone', c.get('phone_number') or '')
            c.setdefault('name', c.get('contact_name') or '')
        return jsonify({'contacts': contacts}), 200
    except Exception as e:
        logger.error(f"Error fetching deleted contacts: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/delete-batch', methods=['POST'])
def level2_delete_batch():
    """Delete a batch and all its contacts"""
    try:
        data = request.json or {}
        batch_name = data.get('batch_name', '').strip()
        
        if not batch_name:
            return jsonify({'error': 'batch_name is required'}), 400
        
        result = get_supabase_client().delete_batch(batch_name)
        
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Failed to delete batch')}), 500
        
        return jsonify({
            'success': True,
            'message': f'Batch "{batch_name}" deleted successfully',
            'deleted_contacts': result.get('deleted_contacts', 0)
        }), 200
    except Exception as e:
        print(f"Error deleting batch: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/merge-batches', methods=['POST'])
def level2_merge_batches():
    """Merge all batches for a project into one batch"""
    try:
        data = request.json or {}
        project_name = data.get('project_name', '').strip()
        target_batch_name = data.get('target_batch_name', '').strip()
        
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        
        # Use project name as default batch name if not provided
        if not target_batch_name:
            target_batch_name = f"{project_name}_Main_Batch"
        
        result = get_supabase_client().merge_duplicate_batches(project_name, target_batch_name)
        
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Failed to merge batches')}), 500
        
        return jsonify({
            'success': True,
            'message': f'Merged {result.get("merged_contacts", 0)} contacts into "{target_batch_name}"',
            'merged_contacts': result.get('merged_contacts', 0),
            'batch_name': target_batch_name
        }), 200
    except Exception as e:
        print(f"Error merging batches: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/level2/delete-duplicate-batches', methods=['POST'])
def level2_delete_duplicate_batches():
    """Delete all duplicate batches for a project, keeping only the main one"""
    try:
        data = request.json or {}
        project_name = data.get('project_name', '').strip()
        
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        
        # Get all batches for this project
        batches = get_supabase_client().get_batches_list(project_name=project_name)
        
        if len(batches) <= 1:
            return jsonify({
                'success': True,
                'message': 'No duplicate batches to delete',
                'deleted': 0
            }), 200
        
        # Keep the batch with most contacts, delete others
        batches.sort(key=lambda x: x.get('contact_count', 0), reverse=True)
        main_batch = batches[0]
        duplicate_batches = batches[1:]
        
        deleted_count = 0
        deleted_contacts = 0
        
        for batch in duplicate_batches:
            result = get_supabase_client().delete_batch(batch.get('batch_name'))
            if result.get('success'):
                deleted_count += 1
                deleted_contacts += result.get('deleted_contacts', 0)
        
        return jsonify({
            'success': True,
            'message': f'Deleted {deleted_count} duplicate batches',
            'deleted_batches': deleted_count,
            'deleted_contacts': deleted_contacts,
            'kept_batch': main_batch.get('batch_name')
        }), 200
    except Exception as e:
        print(f"Error deleting duplicate batches: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Level 3 routes removed - functionality deprecated

@app.route('/api/level2/delete-companies', methods=['POST'])
def level2_delete_companies():
    """Soft-exclude companies (set excluded_at); data stays in DB and shows in Excluded list"""
    try:
        data = request.json
        project_name = data.get('project_name')
        place_ids = data.get('place_ids', [])
        
        # Get contacts from Supabase (prefer batch_name)
        if batch_name:
            contacts = get_supabase_client().get_contacts_for_level3(batch_name=batch_name)
        else:
            contacts = get_supabase_client().get_contacts_for_level3(project_name=project_name)
        
        if not contacts:
            return jsonify({'error': 'No contacts found in Supabase. Please run Level 2 first.'}), 400
        
        # NOTE: Deprecated bulk transfer path – use /api/level3/transfer-one for per-contact progress.
        return jsonify({
            'success': False,
            'error': 'Use /api/level3/transfer-one for per-contact transfer with progress.'
        }), 400
        
    except Exception as e:
        print(f"Error in Level 3 transfer: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/level3/status', methods=['GET'])
def level3_status():
    """Get status of contacts in Supabase for Level 3"""
    try:
        project_name = request.args.get('project_name')
        
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        
        contacts = get_supabase_client().get_contacts_for_level3(project_name=project_name)
        return jsonify({
            'total_contacts': len(contacts),
            'ready_for_transfer': len(contacts) > 0
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/level3/apollo-email-accounts', methods=['GET'])
def level3_apollo_email_accounts():
    """Fetch Apollo email accounts for 'Send from' dropdown when adding contacts to a sequence."""
    try:
        result = apollo_client.get_email_accounts()
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Failed to fetch email accounts'),
                'email_accounts': []
            }), 200
        accounts = result.get('email_accounts', [])
        return jsonify({
            'success': True,
            'email_accounts': [{"id": a.get("id"), "email": a.get("email"), "active": a.get("active", True), "provider_display_name": a.get("provider_display_name")} for a in accounts]
        }), 200
    except Exception as e:
        logger.exception("Apollo email accounts fetch failed")
        return jsonify({'success': False, 'error': str(e), 'email_accounts': []}), 200

@app.route('/api/level3/apollo-users', methods=['GET'])
def level3_apollo_users():
    """Fetch Apollo users (teammates) for optional user_id when adding contacts to a sequence."""
    try:
        page = int(request.args.get('page', 1))
        per_page = min(int(request.args.get('per_page', 50)), 100)
        result = apollo_client.get_users(page=page, per_page=per_page)
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Failed to fetch users'),
                'users': [],
                'pagination': {}
            }), 200
        users = result.get('users', [])
        return jsonify({
            'success': True,
            'users': [{"id": u.get("id"), "email": u.get("email"), "name": u.get("name") or f"{u.get('first_name') or ''} {u.get('last_name') or ''}".strip() or u.get("email")} for u in users],
            'pagination': result.get('pagination', {})
        }), 200
    except Exception as e:
        logger.exception("Apollo users fetch failed")
        return jsonify({'success': False, 'error': str(e), 'users': [], 'pagination': {}}), 200

@app.route('/api/level3/apollo-sequences', methods=['GET'])
def level3_apollo_sequences():
    """Fetch Apollo sequences (emailer campaigns) for dropdown: Add batch to sequence.
    Requires Apollo master API key. Optional q_name to filter by sequence name."""
    try:
        q_name = request.args.get('q_name', '').strip() or None
        page = int(request.args.get('page', 1))
        per_page = min(int(request.args.get('per_page', 20)), 50)
        result = apollo_client.search_sequences(q_name=q_name, page=page, per_page=per_page)
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Failed to fetch sequences'),
                'sequences': []
            }), 200
        return jsonify({
            'success': True,
            'sequences': result.get('sequences', []),
            'pagination': result.get('pagination', {})
        }), 200
    except Exception as e:
        logger.exception("Apollo sequences fetch failed")
        return jsonify({'success': False, 'error': str(e), 'sequences': []}), 200

@app.route('/api/level3/apollo-add-to-sequence', methods=['POST'])
def level3_apollo_add_to_sequence():
    """Add Apollo contacts to an Apollo sequence. Requires Apollo contact IDs and send_email_from_email_account_id."""
    try:
        data = request.json or {}
        sequence_id = (data.get('sequence_id') or '').strip()
        contact_ids = data.get('contact_ids') or []
        if isinstance(contact_ids, str):
            contact_ids = [contact_ids] if contact_ids else []
        contact_ids = [str(c).strip() for c in contact_ids if c]
        send_email_from_email_account_id = (data.get('send_email_from_email_account_id') or '').strip()
        if not sequence_id:
            return jsonify({'success': False, 'error': 'sequence_id is required'}), 400
        if not send_email_from_email_account_id:
            return jsonify({'success': False, 'error': 'send_email_from_email_account_id is required (connected email account ID)'}), 400
        if not contact_ids:
            return jsonify({'success': False, 'error': 'contact_ids is required (list of contact IDs on your outreach platform)'}), 400
        result = apollo_client.add_contacts_to_sequence(
            sequence_id=sequence_id,
            contact_ids=contact_ids,
            send_email_from_email_account_id=send_email_from_email_account_id,
            send_email_from_email_address=data.get('send_email_from_email_address') or None,
            sequence_no_email=bool(data.get('sequence_no_email')),
            sequence_unverified_email=bool(data.get('sequence_unverified_email')),
            sequence_active_in_other_campaigns=bool(data.get('sequence_active_in_other_campaigns')),
            sequence_finished_in_other_campaigns=bool(data.get('sequence_finished_in_other_campaigns')),
            status=data.get('status') or None,
            user_id=data.get('user_id') or None,
        )
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Failed to add contacts to sequence'),
                'contacts': result.get('contacts', []),
                'skipped_contact_ids': result.get('skipped_contact_ids', {})
            }), 200
        return jsonify({
            'success': True,
            'contacts': result.get('contacts', []),
            'skipped_contact_ids': result.get('skipped_contact_ids', {}),
            'emailer_campaign': result.get('emailer_campaign'),
        }), 200
    except Exception as e:
        logger.exception("Apollo add to sequence failed")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/level3/apollo-update-sequence-status', methods=['POST'])
def level3_apollo_update_sequence_status():
    """Update contact status in Apollo sequence: mark_as_finished, remove, or stop."""
    try:
        data = request.json or {}
        emailer_campaign_ids = data.get('emailer_campaign_ids') or data.get('sequence_ids') or []
        if isinstance(emailer_campaign_ids, str):
            emailer_campaign_ids = [emailer_campaign_ids] if emailer_campaign_ids else []
        emailer_campaign_ids = [str(s).strip() for s in emailer_campaign_ids if s]
        contact_ids = data.get('contact_ids') or []
        if isinstance(contact_ids, str):
            contact_ids = [contact_ids] if contact_ids else []
        contact_ids = [str(c).strip() for c in contact_ids if c]
        mode = (data.get('mode') or '').strip()
        if not emailer_campaign_ids:
            return jsonify({'success': False, 'error': 'emailer_campaign_ids (or sequence_ids) is required'}), 400
        if not contact_ids:
            return jsonify({'success': False, 'error': 'contact_ids is required'}), 400
        if not mode:
            return jsonify({'success': False, 'error': 'mode is required (mark_as_finished, remove, or stop)'}), 400
        result = apollo_client.update_contact_status_in_sequence(
            emailer_campaign_ids=emailer_campaign_ids,
            contact_ids=contact_ids,
            mode=mode,
        )
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Failed to update status'),
                'contacts': result.get('contacts', []),
            }), 200
        return jsonify({
            'success': True,
            'contacts': result.get('contacts', []),
            'emailer_campaigns': result.get('emailer_campaigns', []),
            'num_contacts': result.get('num_contacts'),
            'contact_statuses': result.get('contact_statuses', {}),
        }), 200
    except Exception as e:
        logger.exception("Apollo update sequence status failed")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/level3/apollo-bulk-create-contacts', methods=['POST'])
def level3_apollo_bulk_create_contacts():
    """Bulk create contacts in Apollo (max 100 per request). Returns created_contacts and existing_contacts with ids."""
    try:
        data = request.json or {}
        contacts = data.get('contacts') or []
        append_label_names = data.get('append_label_names') or []
        run_dedupe = bool(data.get('run_dedupe', False))
        if not contacts:
            return jsonify({'success': False, 'error': 'contacts array is required', 'created_contacts': [], 'existing_contacts': []}), 400
        if len(contacts) > 100:
            return jsonify({'success': False, 'error': 'Maximum 100 contacts per request', 'created_contacts': [], 'existing_contacts': []}), 400
        result = apollo_client.bulk_create_contacts(contacts=contacts, append_label_names=append_label_names or None, run_dedupe=run_dedupe)
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Bulk create failed'),
                'created_contacts': result.get('created_contacts', []),
                'existing_contacts': result.get('existing_contacts', []),
            }), 200
        return jsonify({
            'success': True,
            'created_contacts': result.get('created_contacts', []),
            'existing_contacts': result.get('existing_contacts', []),
        }), 200
    except Exception as e:
        logger.exception("Apollo bulk create contacts failed")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/level3/apollo-search-contacts', methods=['GET', 'POST'])
def level3_apollo_search_contacts():
    """Search Apollo contacts by keywords (name, email, company, title). Returns contacts with id for add-to-sequence."""
    try:
        if request.method == 'POST':
            data = request.json or {}
            q_keywords = (data.get('q_keywords') or '').strip() or None
            page = int(data.get('page', 1))
            per_page = min(int(data.get('per_page', 25)), 100)
        else:
            q_keywords = request.args.get('q_keywords', '').strip() or None
            page = int(request.args.get('page', 1))
            per_page = min(int(request.args.get('per_page', 25)), 100)
        result = apollo_client.search_contacts(q_keywords=q_keywords, page=page, per_page=per_page)
        if not result.get('success'):
            return jsonify({'success': False, 'error': result.get('error', 'Search failed'), 'contacts': [], 'pagination': {}}), 200
        return jsonify({
            'success': True,
            'contacts': result.get('contacts', []),
            'pagination': result.get('pagination', {}),
        }), 200
    except Exception as e:
        logger.exception("Apollo search contacts failed")
        return jsonify({'success': False, 'error': str(e), 'contacts': [], 'pagination': {}}), 500

@app.route('/api/level3/create-list', methods=['POST'])
def level3_create_list():
    """Create a list in Outreach Platform and return list_id"""
    try:
        data = request.json or {}
        list_name = data.get('list_name', '').strip()
        if not list_name:
            return jsonify({'error': 'list_name is required'}), 400

        result = apollo_client.create_contact_list(list_name)
        if result.get('success'):
            return jsonify({'success': True, 'list_id': result.get('list_id')}), 200
        return jsonify({'error': result.get('error', 'Failed to create list')}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/level3/contacts', methods=['GET'])
def level3_contacts():
    """Return contact list for a batch (for preview and transfer).
    Pass designation from Level 2 so the count matches (e.g. 30 in L2 = 30 in L3).
    """
    try:
        batch_name = request.args.get('batch_name')
        designation = (request.args.get('designation') or '').strip() or None
        if not batch_name:
            return jsonify({'error': 'batch_name is required'}), 400

        contacts = get_supabase_client().get_contacts_for_level3(batch_name=batch_name, designation=designation)
        
        # Safety filter: Exclude generic employee titles (in case old batches weren't filtered)
        excluded_titles = ['employee', 'staff', 'worker', 'member', 'personnel']
        filtered_contacts = []
        for c in contacts:
            title_lower = (c.get('title') or '').lower().strip()
            # Skip if title is empty or is a generic employee title
            if title_lower and title_lower not in excluded_titles:
                filtered_contacts.append(c)
            elif not title_lower:
                # Include contacts without title (might be valid)
                filtered_contacts.append(c)
        
        # Return contacts with proper formatting
        minimal = []
        for c in filtered_contacts:
            # Use title field directly - same as Level 2 saved
            display_title = (c.get('title') or '').strip()
            
            minimal.append({
                'id': c.get('id'),
                'name': c.get('contact_name', '') or c.get('name', ''),
                'email': c.get('email', ''),
                'company_name': c.get('company_name', ''),
                'title': display_title,
                'contact_type': c.get('contact_type', ''),
                'industry': (c.get('industry') or '').strip()
            })
        
        excluded_count = len(contacts) - len(minimal)
        if excluded_count > 0:
            print(f"⚠️  Level 3: Excluded {excluded_count} employee contacts (safety filter)")
        print(f"✅ Level 3: Returning {len(minimal)} contacts (filtered from {len(contacts)} total)")
        return jsonify({'success': True, 'contacts': minimal, 'count': len(minimal)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/level3/ensure-companies', methods=['POST'])
def level3_ensure_companies():
    """Create unique companies (accounts) in Apollo Companies for the current batch before transfer."""
    try:
        data = request.json or {}
        batch_name = data.get('batch_name') or request.args.get('batch_name')
        if not batch_name:
            return jsonify({'error': 'batch_name is required'}), 400

        contacts = get_supabase_client().get_contacts_for_level3(batch_name=batch_name, designation=None)
        if not contacts:
            return jsonify({'success': True, 'companies_created': 0, 'message': 'No contacts in batch'}), 200

        # Unique companies by (name, domain)
        seen = set()
        companies = []
        for c in contacts:
            name = (c.get('company_name') or '').strip()
            website = (c.get('company_website') or '').strip()
            domain = apollo_client.extract_domain(website) if website else ''
            key = (name or '', domain or '')
            if key in seen or (not name and not domain):
                continue
            seen.add(key)
            companies.append({
                'name': name or domain or 'Unknown',
                'domain': domain,
                'phone': (c.get('company_phone') or '').strip(),
                'raw_address': (c.get('company_address') or '').strip()
            })

        created = 0
        errors = []
        for co in companies:
            r = apollo_client.create_account(
                name=co['name'],
                domain=co['domain'],
                phone=co['phone'],
                raw_address=co['raw_address']
            )
            if r.get('success'):
                created += 1
            else:
                errors.append(f"{co['name']}: {r.get('error', '')}")

        return jsonify({
            'success': True,
            'companies_created': created,
            'companies_total': len(companies),
            'errors': errors[:10] if errors else None
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Level 3 routes removed - functionality deprecated

@app.route('/api/level2/restore-companies', methods=['POST'])
def level2_restore_companies():
    """Restore excluded companies (clear excluded_at) so they show in the main list again."""
    try:
        data = request.json
        project_name = data.get('project_name')
        place_ids = data.get('place_ids', [])
        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400
        if not place_ids:
            return jsonify({'error': 'No place_ids provided'}), 400
        result = get_supabase_client().restore_level1_companies(project_name=project_name, identifiers=place_ids)
        if result.get('success'):
            return jsonify({'success': True, 'restored': result.get('restored', 0), 'message': f'Restored {result.get("restored", 0)} companies'}), 200
        return jsonify({'error': result.get('error', 'Failed to restore')}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
def export():
    """Export Level 1 company data to Excel file (Places API only - no contacts)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from flask import send_file
        import io
        
        data = request.json
        companies = data.get('companies', [])
        project_name = data.get('project_name', 'companies')
        
        if not companies:
            return jsonify({'error': 'No companies to export'}), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Companies"
        
        # Headers
        headers = ['Company Name', 'Website', 'Phone', 'Address', 'Industry', 'PIN Code', 'Place ID', 'Place Type']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for company in companies:
            ws.append([
                company.get('company_name', ''),
                company.get('website', ''),
                company.get('phone', ''),
                company.get('address', ''),
                company.get('industry', ''),
                company.get('pin_code', ''),
                company.get('place_id', ''),
                company.get('place_type', '')
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename
        safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_', '/')).strip()[:50]
        filename = f"{safe_project_name}_companies.xlsx" if safe_project_name else "companies.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-contacts', methods=['POST'])
def export_contacts():
    """Export Level 3 contacts data to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from flask import send_file
        import io
        
        data = request.json
        contacts = data.get('contacts', [])
        batch_name = data.get('batch_name', 'contacts')
        
        if not contacts:
            return jsonify({'error': 'No contacts to export'}), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Headers
        headers = ['Name', 'Title', 'Company Name', 'Email', 'Phone', 'LinkedIn URL', 'Company Website', 'Company Address']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for contact in contacts:
            ws.append([
                contact.get('name') or contact.get('contact_name', ''),
                contact.get('title') or contact.get('contact_type', ''),
                contact.get('company_name', ''),
                contact.get('email', ''),
                contact.get('phone', ''),
                contact.get('linkedin_url', ''),
                contact.get('company_website', ''),
                contact.get('company_address', '')
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename
        safe_batch_name = "".join(c for c in batch_name if c.isalnum() or c in (' ', '-', '_', '/')).strip()[:50]
        filename = f"{safe_batch_name}_contacts.xlsx" if safe_batch_name else "contacts.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting contacts to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/apollo/webhook', methods=['POST'])
def apollo_webhook():
    """
    Webhook endpoint to receive phone numbers from Enrichment Service
    Service sends phone numbers via webhook after enrichment request
    """
    try:
        data = request.json or {}
        
        # Enrichment service webhook payload structure
        person = data.get('person', {}) or data.get('data', {}).get('person', {})
        if not person:
            print("⚠️  Enrichment webhook: No person data in payload")
            return jsonify({'success': False, 'error': 'No person data'}), 400
        
        person_id = person.get('id') or data.get('person_id')
        email = person.get('email', '')
        phone_numbers = person.get('phone_numbers', [])
        
        print(f"📞 Enrichment webhook received for person_id: {person_id}, email: {email}")
        print(f"📞 Phone numbers: {phone_numbers}")
        
        # Extract phone number
        phone = ''
        if phone_numbers and len(phone_numbers) > 0:
            phone_obj = phone_numbers[0]
            phone = (
                phone_obj.get('raw_number', '') or
                phone_obj.get('sanitized_number', '') or
                phone_obj.get('number', '') or
                phone_obj.get('phone', '')
            )
        
        if phone:
            print(f"✅ Extracted phone from webhook: {phone}")
            # Update contact in database with phone number
            # Match by email (most reliable)
            if email:
                try:
                    # Update contacts with this email in level2_contacts table
                    supabase = get_supabase_client()
                    # Update all contacts with matching email
                    result = supabase.client.table('level2_contacts').update({
                        'phone_number': phone
                    }).eq('email', email).execute()
                    
                    if result.data:
                        print(f"✅ Updated {len(result.data)} contact(s) with phone number: {phone}")
                    else:
                        print(f"⚠️  No contacts found with email: {email}")
                except Exception as e:
                    print(f"⚠️  Could not update contact with phone: {str(e)}")
        else:
            print(f"⚠️  No phone number in webhook payload")
        
        return jsonify({'success': True, 'phone': phone}), 200
        
    except Exception as e:
        print(f"❌ Error processing enrichment webhook: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Vercel serverless handler (required for Vercel Python runtime)
# Vercel expects the app to be directly callable
# The @vercel/python builder automatically wraps Flask apps
# Updated: 2026-01-24 - Multiple employee range selection feature added

if __name__ == '__main__':
    Config.validate()
    port = int(os.getenv('PORT', 5002))
    app.run(debug=False, host='0.0.0.0', port=port)

